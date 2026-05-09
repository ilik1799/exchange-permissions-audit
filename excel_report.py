"""
Excel-отчёт с цветовой дифференциацией.

Красный — Полный доступ
Жёлтый  — Отправить как
Зелёный — Отправить от имени
"""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

PERM_COLORS = {
    "Полный доступ": PatternFill("solid", fgColor="FFDDD9"),
    "Отправить как": PatternFill("solid", fgColor="FFF3CD"),
    "Отправить от имени": PatternFill("solid", fgColor="D4EDDA"),
}

COLUMNS = [
    ("№", 5),
    ("Почтовый ящик", 35),
    ("Описание", 25),
    ("Право доступа", 22),
    ("Кому предоставлено", 30),
    ("Детали", 55),
]


def export(
    report: list[dict],
    target_user: str,
    total_mailboxes: int,
    threads: int,
    elapsed: str,
    output_path: str,
) -> None:
    """Формирует Excel и сохраняет по пути."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None

    ws.title = "Exchange Permissions"

    hdr_font = Font(
        name="Calibri",
        bold=True,
        size=11,
        color="FFFFFF",
    )
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    hdr_align = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    c_font = Font(name="Calibri", size=10)
    c_align = Alignment(vertical="top", wrap_text=True)
    brd = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for ci, (title, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=title)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = brd
        col_letter = cell.column_letter
        ws.column_dimensions[col_letter].width = width

    for ri, entry in enumerate(report, 2):
        perm = entry.get("Permission", "")
        data = [
            ri - 1,
            entry.get("Mailbox", ""),
            entry.get("Description", ""),
            perm,
            entry.get("GrantedTo", ""),
            entry.get("Detail", ""),
        ]
        fill = PERM_COLORS.get(perm)
        for ci, val in enumerate(data, 1):
            cell = ws.cell(
                row=ri,
                column=ci,
                value=val,
            )
            cell.font = c_font
            cell.alignment = c_align
            cell.border = brd
            if fill:
                cell.fill = fill

    ws.auto_filter.ref = ws.dimensions  # type: ignore
    ws.freeze_panes = "A2"  # type: ignore

    r = len(report) + 3
    meta = [
        f"Отчёт для: {target_user}",
        f"Проверено ящиков: {total_mailboxes}",
        f"Параллельных сессий: {threads}",
        f"Время выполнения: {elapsed}",
        f"Дата: {datetime.now():%d.%m.%Y %H:%M}",
    ]
    for i, text in enumerate(meta):
        c = ws.cell(row=r + i, column=1, value=text)
        c.font = Font(italic=True, size=9)

    wb.save(output_path)
